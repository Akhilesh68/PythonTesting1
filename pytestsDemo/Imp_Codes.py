# n=int(input("Enter Num: "))
# x=0
# y=1
# z=0
# while (z<=n):
#     print(z)
#     x=y
#     y=z
#     z=x+y
# print(z)

import openpyxl

workbook = openpyxl.load_workbook("C:\\Users\Akhilesh\\Desktop\\ETL_Testing_Test_Cases.xlsx")
sheet= workbook.active
print(sheet)
sheet=workbook["ETL_Priority_Severity"]
print(sheet)
max_rows = sheet.max_row
max_columns=sheet.max_column
print(max_rows,max_columns)



